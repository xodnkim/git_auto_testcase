# utils/exporter.py
import pandas as pd
import io
import re
import markdown
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

def generate_html_report(md_text):
    """마크다운에서 CSV 블록을 완전히 제거하고 순수 보고서(1~3번)만 HTML로 변환합니다."""
    clean_text = re.sub(r'```csv\n(.*?)\n```', '', md_text, flags=re.DOTALL | re.IGNORECASE).strip()
    
    html_body = markdown.markdown(clean_text, extensions=['tables'])
    
    html_template = f"""
    <!DOCTYPE html>
    <html lang="ko">
    <head>
        <meta charset="UTF-8">
        <title>QA 리스크 분석 보고서</title>
        <style>
            body {{ font-family: 'Malgun Gothic', sans-serif; line-height: 1.6; padding: 40px; max-width: 900px; margin: auto; color: #333; }}
            h1, h2, h3 {{ color: #2C3E50; border-bottom: 1px solid #eee; padding-bottom: 10px; }}
            ul {{ margin-bottom: 20px; }}
        </style>
    </head>
    <body>
        <h1>🛡️ QA 리스크 분석 보고서</h1>
        <hr>
        {html_body}
    </body>
    </html>
    """
    return html_template.encode('utf-8')

def generate_tc_excel(md_text):
    """CSV 블록을 추출하여 완벽한 실무 서식이 적용된 엑셀 파일로 변환합니다."""
    csv_match = re.search(r'```csv\n(.*?)\n```', md_text, re.DOTALL | re.IGNORECASE)
    if not csv_match: return None 
        
    csv_data = csv_match.group(1).strip()
    
    try:
        df = pd.read_csv(io.StringIO(csv_data))
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Test Cases')
            worksheet = writer.sheets['Test Cases']
            
            # 🎨 엑셀 서식 지정 세팅
            header_fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")
            header_font = Font(bold=True)
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

            # 1. 헤더 스타일 적용
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.border = thin_border
                cell.alignment = center_align

            # 2. 컬럼별 맞춤 너비 설정 (TC_No부터 Result까지)
            col_widths = {'A': 12, 'B': 15, 'C': 15, 'D': 15, 'E': 30, 'F': 25, 'G': 25, 'H': 25, 'I': 10}
            for col_letter, width in col_widths.items():
                worksheet.column_dimensions[col_letter].width = width

            # 3. 데이터 셀 스타일 적용 (상세, 조건, 동작, 결과는 좌측 정렬)
            for row in worksheet.iter_rows(min_row=2, max_col=9, max_row=worksheet.max_row):
                for cell in row:
                    cell.border = thin_border
                    if cell.column_letter in ['E', 'F', 'G', 'H']:
                        cell.alignment = left_align
                    else:
                        cell.alignment = center_align
                        
        return output.getvalue()
    except Exception as e:
        print(f"엑셀 변환 에러: {e}")
        return None